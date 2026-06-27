import os
from flask import render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user
from sqlalchemy.exc import IntegrityError  
from sqlalchemy import or_
from openpyxl import load_workbook         
from datetime import datetime
from . import gestion_bp
from ..models import Paralelo, Usuario, Inscripcion, Rol
from ..extensions import db, bcrypt
import re # IMPORTANTE: Nuestra herramienta "colador" para limpiar letras y basura de los inputs


@gestion_bp.route('/paralelo/<int:id>/estudiantes', methods=['GET', 'POST'])
@login_required
def estudiantes(id):
    paralelo = Paralelo.query.get_or_404(id)
    if paralelo.auxiliar_id != current_user.id:
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('gestion.paralelos'))

    if request.method == 'POST':
        nombres = request.form.get('nombres')
        apellidos = request.form.get('apellidos')
        
        # --- Sanitización: El famoso 'colador' ---
        # No confiamos en el form. Dejamos solo números puros para no tener líos con los 'LP' o espacios
        ci_crudo = request.form.get('ci', '').strip()
        ci = re.sub(r'\D', '', ci_crudo) 
        
        ru_crudo = request.form.get('ru', '').strip()
        ru = re.sub(r'\D', '', ru_crudo) if ru_crudo else None

        estudiante = Usuario.query.filter_by(ci=ci).first()
        
        # Si el estudiante es nuevo, le armamos su cuenta al vuelo
        if not estudiante:
            rol_estudiante = Rol.query.filter_by(nombre='Estudiante').first()
            hashed_pw = bcrypt.generate_password_hash(ci).decode('utf-8') # Por defecto, C.I. como password
            estudiante = Usuario(
                nombres=nombres.upper(), 
                apellidos=apellidos.upper(), 
                ci=ci, 
                ru=ru, 
                username=ci, # Por defecto, C.I. como username
                password_hash=hashed_pw, 
                rol_id=rol_estudiante.id,
                estado=True
            )
            db.session.add(estudiante)
            
            # Savepoint: Si esto explota por un duplicado fantasma, lo atajamos sin tumbar el servidor
            try:
                db.session.flush()
            except IntegrityError:
                db.session.rollback()
                flash('Error: El C.I. o el R.U. ingresado ya le pertenece a otra cuenta en el sistema.', 'danger')
                return redirect(url_for('gestion.estudiantes', id=id))

        # Control de inscripción manual garantizando el estado activo
        inscripcion_previa = Inscripcion.query.filter_by(estudiante_id=estudiante.id, paralelo_id=id).first()
        
        if sig_insc := inscripcion_previa:
            if not sig_insc.estado:
                sig_insc.estado = True # Si estaba dado de baja, lo revivimos mágicamente
                flash('Inscripción del estudiante reactivada.', 'success')
            else:
                flash('El estudiante ya está inscrito en este paralelo.', 'warning')
        else:
            # Inscripción limpiecita desde cero
            nueva_inscripcion = Inscripcion(estudiante_id=estudiante.id, paralelo_id=id, estado=True)
            db.session.add(nueva_inscripcion)
            flash('Estudiante inscrito exitosamente.', 'success')
            
        db.session.commit()
        return redirect(url_for('gestion.estudiantes', id=id))

    # Consultamos y ordenamos alfabéticamente para que la vista del Auxi sea impecable
    inscripciones_bd = Inscripcion.query.filter_by(paralelo_id=id, estado=True).all()
    inscripciones = sorted(inscripciones_bd, key=lambda i: (i.estudiante.apellidos, i.estudiante.nombres))
    return render_template('gestion/estudiantes.html', paralelo=paralelo, inscripciones=inscripciones)


@gestion_bp.route('/estudiante/<int:id>/editar', methods=['POST'])
@login_required
def editar_estudiante(id):
    estudiante = Usuario.query.get_or_404(id)
    paralelo_id = request.form.get('paralelo_id')
    
    estudiante.nombres = request.form.get('nombres').upper()
    estudiante.apellidos = request.form.get('apellidos').upper()
    
    # Sanitización al Editar: Volvemos a aplicar el colador por si el Auxi metió una letra sin querer
    ci_crudo = request.form.get('ci', '').strip()
    estudiante.ci = re.sub(r'\D', '', ci_crudo)
    
    ru_crudo = request.form.get('ru', '').strip()
    estudiante.ru = re.sub(r'\D', '', ru_crudo) if ru_crudo else None
    
    try:
        db.session.commit()
        flash('Datos del estudiante actualizados correctamente.', 'success')
    except IntegrityError:
        db.session.rollback() # Evitamos colisiones si le intenta poner un R.U. que ya existe
        flash('Error de duplicidad: El C.I. o el R.U. que intentas asignar ya le pertenece a otro estudiante.', 'danger')
        
    return redirect(url_for('gestion.estudiantes', id=paralelo_id))


@gestion_bp.route('/inscripcion/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_inscripcion(id):
    inscripcion = Inscripcion.query.get_or_404(id)
    paralelo_id = inscripcion.paralelo_id
    if inscripcion.paralelo.auxiliar_id == current_user.id:
        # Baja lógica: Nunca hacemos DELETE físico para no huérfanizar las notas
        inscripcion.estado = False
        db.session.commit()
        flash('Estudiante dado de baja del paralelo.', 'info')
    return redirect(url_for('gestion.estudiantes', id=paralelo_id))


# --- MOTOR BLINDADO: IMPORTACIÓN MASIVA DESDE EXCEL ---
# Con doble validación anti-colisiones y savepoints anidados
@gestion_bp.route('/paralelo/<int:id>/importar_estudiantes', methods=['POST'])
@login_required
def importar_estudiantes(id):
    paralelo = Paralelo.query.get_or_404(id)
    
    if paralelo.auxiliar_id != current_user.id:
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('gestion.paralelos'))

    if 'archivo_excel' not in request.files:
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('gestion.estudiantes', id=id))

    file = request.files['archivo_excel']
    if file.filename == '':
        flash('El archivo no tiene un nombre válido.', 'danger')
        return redirect(url_for('gestion.estudiantes', id=id))

    if file and file.filename.endswith(('.xlsx', '.xlsm')):
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active 

            alumnos_inscritos = 0
            alumnos_existentes_en_paralelo = 0
            errores_omitidos = 0 # Contador para saber cuántos chocaron en el intento
            rol_estudiante = Rol.query.filter_by(nombre='Estudiante').first()

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue

                nombre_completo_crudo = str(row[0]).strip()
                
                # --- Sanitización extrema del EXCEL ---
                # Las listas oficiales siempre traen basura. Aquí lo dejamos impecable.
                ci_crudo = str(row[1]).strip()
                ci = re.sub(r'\D', '', ci_crudo) 
                
                ru_crudo = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
                ru = re.sub(r'\D', '', ru_crudo) if ru_crudo else None

                if not nombre_completo_crudo or not ci:
                    continue

                # Inteligencia artificial rústica para separar Apellidos y Nombres
                if ',' in nombre_completo_crudo:
                    partes = nombre_completo_crudo.split(',')
                    apellidos = partes[0].strip().upper()
                    nombres = partes[1].strip().upper()
                else:
                    palabras = nombre_completo_crudo.split()
                    if len(palabras) >= 4:
                        apellidos = f"{palabras[0]} {palabras[1]}".upper()
                        nombres = " ".join(palabras[2:]).upper()
                    elif len(palabras) == 3:
                        apellidos = f"{palabras[0]} {palabras[1]}".upper()
                        nombres = palabras[2].upper()
                    elif len(palabras) == 2:
                        apellidos = palabras[0].upper()
                        nombres = palabras[1].upper()
                    else:
                        apellidos = "SIN APELLIDO"
                        nombres = palabras[0].upper()

                # Búsqueda expansiva: ¿Ya lo tenemos en el sistema?
                estudiante = Usuario.query.filter(or_(Usuario.ci == ci, Usuario.username == ci)).first()

                if not estudiante:
                    # SAVEPOINT 1: Intentamos crearlo. Si algo falla (ej. RU repetido), lo omitimos y seguimos
                    try:
                        with db.session.begin_nested():
                            hashed_pw = bcrypt.generate_password_hash(ci).decode('utf-8')
                            estudiante = Usuario(
                                nombres=nombres, 
                                apellidos=apellidos, 
                                ci=ci, 
                                ru=ru, 
                                username=ci, 
                                password_hash=hashed_pw, 
                                rol_id=rol_estudiante.id,
                                estado=True
                            )
                            db.session.add(estudiante)
                    except IntegrityError:
                        errores_omitidos += 1
                        continue 

                if estudiante: 
                    inscripcion_existente = Inscripcion.query.filter_by(
                        paralelo_id=paralelo.id, 
                        estudiante_id=estudiante.id
                    ).first()

                    # SAVEPOINT 2: Matriculación segura
                    if not inscripcion_existente:
                        # ... dentro del if not inscripcion_existente:
                        try:
                            with db.session.begin_nested():
                                # Usamos datetime.utcnow() que es más compatible con bases de datos en la nube
                                nueva_inscripcion = Inscripcion(
                                    paralelo_id=paralelo.id,
                                    estudiante_id=estudiante.id,
                                    fecha_inscripcion=datetime.utcnow(), 
                                    estado=True
                                )
                                db.session.add(nueva_inscripcion)
                                alumnos_inscritos += 1
                        # ...
                        except IntegrityError:
                            errores_omitidos += 1
                            continue
                    else:
                        if not inscripcion_existente.estado:
                            try:
                                with db.session.begin_nested():
                                    inscripcion_existente.estado = True
                                    alumnos_inscritos += 1
                            except IntegrityError:
                                errores_omitidos += 1
                                continue
                        else:
                            alumnos_existentes_en_paralelo += 1

            # COMMIT MÁGICO: Guardamos a todos los sobrevivientes de un solo golpe
            db.session.commit()
            
            # Reporte detallado para saber exactamente qué pasó en las tripas del proceso
            if alumnos_inscritos > 0:
                flash(f'Éxito: Se matricularon {alumnos_inscritos} estudiantes correctamente.', 'success')
            if alumnos_existentes_en_paralelo > 0:
                flash(f'Aviso: {alumnos_existentes_en_paralelo} alumnos del Excel ya estaban inscritos.', 'info')
            if errores_omitidos > 0:
                flash(f'Advertencia: Se omitieron {errores_omitidos} filas por conflicto de datos (Ej: R.U. duplicado).', 'warning')
            if alumnos_inscritos == 0 and alumnos_existentes_en_paralelo == 0 and errores_omitidos == 0:
                flash('No se encontraron registros válidos para importar en el archivo.', 'danger')

        except Exception as e:
            db.session.rollback()
            flash(f'Error al procesar el archivo Excel: Asegúrate de usar el formato correcto.', 'danger')
            print(f"Error interno: {str(e)}") # Para debuggear en consola sin asustar al usuario
    else:
        flash('Formato no permitido. Sube un archivo Excel válido (.xlsx).', 'danger')

    return redirect(url_for('gestion.estudiantes', id=id))


# --- EL SALVAVIDAS DEL AUXI ---
# Para cuando los alumnos se olvidan su contraseña o cambian su username y se quedan afuera
@gestion_bp.route('/estudiante/<int:estudiante_id>/restablecer_credenciales', methods=['POST'])
@login_required
def restablecer_credenciales(estudiante_id):
    estudiante = Usuario.query.get_or_404(estudiante_id)
    
    # Reseteo forzoso a modo fábrica (C.I.)
    estudiante.username = estudiante.ci
    estudiante.password_hash = bcrypt.generate_password_hash(estudiante.ci).decode('utf-8')
    db.session.commit()
    
    flash(f'Credenciales de {estudiante.nombres} restablecidas con éxito. Usuario y Contraseña ahora son su C.I.', 'success')
    return redirect(request.referrer or url_for('gestion.paralelos'))