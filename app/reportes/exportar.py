import io
from flask import send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from . import reportes_bp
from ..models import Paralelo, ParametroEvaluacion, Actividad, Calificacion

@reportes_bp.route('/paralelo/<int:id>/exportar_excel')
@login_required
def exportar_excel(id):
    paralelo = Paralelo.query.get_or_404(id)
    if paralelo.auxiliar_id != current_user.id:
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('reportes.selector'))

    parametros = ParametroEvaluacion.query.filter_by(paralelo_id=id, estado=True).order_by(ParametroEvaluacion.id).all()
    
    inscripciones_activas = [insc for insc in paralelo.inscripciones if insc.estado]
    estudiantes = sorted([insc.estudiante for insc in inscripciones_activas], key=lambda e: (e.apellidos, e.nombres))
    
    parametros_evaluacion = [p for p in parametros if p.tipo != 'liberacion']
    param_liberacion = next((p for p in parametros if p.tipo == 'liberacion'), None)

    matriz = []
    for estudiante in estudiantes:
        fila = {
            'estudiante': estudiante,
            'notas_regulares': {},    
            'detalle_actividades': {},
            'nota_semestre': 0.0,
            'nota_liberacion': None,
            'nota_final': 0.0
        }

        nota_base = 0.0
        nota_extra = 0.0

        for param in parametros_evaluacion:
            actividades = Actividad.query.filter_by(parametro_id=param.id, estado=True).all()
            suma_puntajes_100 = 0.0
            
            for act in actividades:
                calif = Calificacion.query.filter_by(actividad_id=act.id, estudiante_id=estudiante.id, estado=True).first()
                puntaje = calif.puntaje if calif else 0.0
                fila['detalle_actividades'][act.id] = puntaje
                suma_puntajes_100 += puntaje

            if len(actividades) > 0:
                promedio_100 = suma_puntajes_100 / len(actividades)
                nota_convertida = (promedio_100 / 100.0) * param.ponderacion
            else:
                nota_convertida = 0.0
                
            fila['notas_regulares'][param.id] = round(nota_convertida, 2)
            
            if param.tipo == 'extra':
                nota_extra += nota_convertida
            else:
                nota_base += nota_convertida

        nota_semestre_bruta = nota_base + nota_extra
        fila['nota_semestre'] = round(min(nota_semestre_bruta, paralelo.nota_maxima), 2)

        if param_liberacion:
            actividades_lib = Actividad.query.filter_by(parametro_id=param_liberacion.id, estado=True).first()
            if actividades_lib:
                calif_lib = Calificacion.query.filter_by(actividad_id=actividades_lib.id, estudiante_id=estudiante.id, estado=True).first()
                puntaje_lib_100 = calif_lib.puntaje if calif_lib else 0.0
                fila['detalle_actividades'][actividades_lib.id] = puntaje_lib_100
                
                nota_liberacion_convertida = (puntaje_lib_100 / 100.0) * param_liberacion.ponderacion
                fila['nota_liberacion'] = round(nota_liberacion_convertida, 2)
            else:
                fila['nota_liberacion'] = 0.0
                
            if fila['nota_liberacion'] is not None and fila['nota_liberacion'] > 0:
                if param_liberacion.modo_liberacion == 'reemplazo':
                    fila['nota_final'] = fila['nota_liberacion']
                else: 
                    fila['nota_final'] = max(fila['nota_semestre'], fila['nota_liberacion'])
            else:
                fila['nota_final'] = fila['nota_semestre']
        else:
            fila['nota_final'] = fila['nota_semestre']
            
        # REDONDEO ACADÉMICO PARA LA NOTA FINAL EN EXCEL
        fila['nota_final'] = int(fila['nota_final'] + 0.5)
        
        matriz.append(fila)

    # 2. Configurar OpenPyXL con Diseño Premium
    wb = Workbook()
    
    font_bold = Font(bold=True)
    font_title = Font(bold=True, size=14, color="FFFFFF")
    font_white = Font(bold=True, color="FFFFFF")
    font_total_cat = Font(bold=True, color="0D47A1") 
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    fill_primary = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid") 
    fill_secondary = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid") 
    fill_zebra = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid") 
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    fill_total_cat = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_semestre = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    fill_liberacion = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid") 
    fill_final = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    fill_final_header = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    
    thin = Side(style='thin', color="BFBFBF")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    def aplicar_borde(celda):
        celda.border = border_all

    # =========================================================================
    # HOJA 1: SÁBANA OFICIAL
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Sábana Oficial"
    
    total_cols_ws1 = 5 + len(parametros_evaluacion) + (1 if param_liberacion else 0) + 2 
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols_ws1)
    ws1['A1'] = f"CENTRALIZADOR OFICIAL - {paralelo.materia.nombre.upper()} ({paralelo.nombre})"
    ws1['A1'].font = font_title
    ws1['A1'].alignment = align_center
    ws1['A1'].fill = fill_primary

    headers_ws1 = ["N°", "Apellidos", "Nombres", "C.I.", "R.U."]
    for param in parametros_evaluacion:
        headers_ws1.append(f"{param.nombre_parametro}\n({param.ponderacion} pts)")
    headers_ws1.append("Semestre")
    if param_liberacion:
        headers_ws1.append("Liberación")
    headers_ws1.append(f"NOTA FINAL\n(Máx. {paralelo.nota_maxima})")

    ws1.append(headers_ws1)
    
    for col_num, cell in enumerate(ws1[2], 1):
        cell.font = font_bold
        cell.alignment = align_center
        aplicar_borde(cell)
        if col_num == total_cols_ws1:
            cell.fill = fill_final_header
            cell.font = font_white
        else:
            cell.fill = fill_secondary

    for i, fila in enumerate(matriz, 1):
        row = [
            i, 
            fila['estudiante'].apellidos, 
            fila['estudiante'].nombres, 
            fila['estudiante'].ci, 
            fila['estudiante'].ru if fila['estudiante'].ru else "-"
        ]
        for param in parametros_evaluacion:
            row.append(fila['notas_regulares'][param.id])
        
        row.append(fila['nota_semestre'])
        if param_liberacion:
            nota_lib = fila['nota_liberacion'] if fila['nota_liberacion'] and fila['nota_liberacion'] > 0 else "-"
            row.append(nota_lib)
        row.append(fila['nota_final'])
        
        ws1.append(row)
        
        current_row = i + 2
        bg_color = fill_zebra if i % 2 == 0 else fill_white
        
        for col_idx in range(1, total_cols_ws1 + 1):
            celda = ws1.cell(row=current_row, column=col_idx)
            aplicar_borde(celda)
            if col_idx in [2, 3]: 
                celda.alignment = align_left
            else:
                celda.alignment = align_center
            
            if col_idx == total_cols_ws1: 
                celda.fill = fill_final
                celda.font = font_bold
            elif param_liberacion and col_idx == total_cols_ws1 - 1:
                celda.fill = fill_liberacion
                celda.font = font_bold
            elif (param_liberacion and col_idx == total_cols_ws1 - 2) or (not param_liberacion and col_idx == total_cols_ws1 - 1):
                celda.fill = fill_semestre
            else:
                celda.fill = bg_color

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 15
    for c in range(6, total_cols_ws1 + 1):
        ws1.column_dimensions[chr(64 + c)].width = 15

    # =========================================================================
    # HOJA 2: DESGLOSE DE AUDITORÍA
    # =========================================================================
    ws2 = wb.create_sheet(title="Desglose de Auditoría")
    ws2.freeze_panes = "F3" 

    titulos_fijos = ["N°", "Apellidos", "Nombres", "C.I.", "R.U."]
    for idx, titulo in enumerate(titulos_fijos, 1):
        celda = ws2.cell(row=1, column=idx, value=titulo)
        celda.font = font_white
        celda.fill = fill_primary
        celda.alignment = align_center
        aplicar_borde(celda)
    
    current_col = 6
    for param in parametros_evaluacion:
        largo = len(param.actividades) + 1 
        ws2.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + largo - 1)
        celda = ws2.cell(row=1, column=current_col, value=param.nombre_parametro.upper())
        celda.alignment = align_center
        celda.font = font_bold
        celda.fill = fill_secondary
        aplicar_borde(celda)
        current_col += largo
            
    if param_liberacion:
        largo = len(param_liberacion.actividades) + 1 
        ws2.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + largo - 1)
        celda = ws2.cell(row=1, column=current_col, value="EXAMEN LIBERACIÓN")
        celda.alignment = align_center
        celda.font = font_white
        celda.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        aplicar_borde(celda)
        current_col += largo

    cols_consolidado = 3 if param_liberacion else 2
    ws2.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + cols_consolidado - 1)
    celda = ws2.cell(row=1, column=current_col, value="CONSOLIDADO FINAL")
    celda.alignment = align_center
    celda.font = font_white
    celda.fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
    aplicar_borde(celda)

    for c in range(1, 6):
        aplicar_borde(ws2.cell(row=2, column=c)) 
        
    current_col = 6
    for param in parametros_evaluacion:
        for act in param.actividades:
            fecha_str = act.fecha.strftime('%d/%m')
            celda = ws2.cell(row=2, column=current_col, value=f"{act.titulo}\n({fecha_str})\n[Sobre 100]")
            celda.alignment = align_center
            celda.font = font_bold
            celda.fill = fill_zebra
            aplicar_borde(celda)
            ws2.column_dimensions[celda.column_letter].width = 16 
            current_col += 1
        
        celda = ws2.cell(row=2, column=current_col, value=f"Total Convertido\n({param.ponderacion} pts)")
        celda.alignment = align_center
        celda.font = font_total_cat
        celda.fill = fill_total_cat
        aplicar_borde(celda)
        ws2.column_dimensions[celda.column_letter].width = 16
        current_col += 1

    if param_liberacion:
        for act in param_liberacion.actividades:
            fecha_str = act.fecha.strftime('%d/%m')
            celda = ws2.cell(row=2, column=current_col, value=f"{act.titulo}\n({fecha_str})\n[Sobre 100]")
            celda.alignment = align_center
            celda.font = font_bold
            celda.fill = fill_zebra
            aplicar_borde(celda)
            ws2.column_dimensions[celda.column_letter].width = 18
            current_col += 1
            
        celda = ws2.cell(row=2, column=current_col, value=f"Total Lib.\n({param_liberacion.ponderacion} pts)")
        celda.alignment = align_center
        celda.font = font_bold
        celda.fill = fill_liberacion
        aplicar_borde(celda)
        ws2.column_dimensions[celda.column_letter].width = 14
        current_col += 1

    titulos_finales = ["Semestre", "Liberación", "NOTA FINAL"] if param_liberacion else ["Semestre", "NOTA FINAL"]
    for idx, titulo in enumerate(titulos_finales):
        celda = ws2.cell(row=2, column=current_col, value=titulo)
        celda.alignment = align_center
        celda.font = font_bold
        if titulo == "Semestre": celda.fill = fill_semestre
        elif titulo == "Liberación": celda.fill = fill_liberacion
        elif titulo == "NOTA FINAL": 
            celda.fill = fill_final_header
            celda.font = font_white
        aplicar_borde(celda)
        ws2.column_dimensions[celda.column_letter].width = 15
        current_col += 1

    for i, fila in enumerate(matriz, 3):
        bg_color = fill_zebra if i % 2 != 0 else fill_white
        
        datos_identidad = [
            i - 2, 
            fila['estudiante'].apellidos, 
            fila['estudiante'].nombres, 
            fila['estudiante'].ci, 
            fila['estudiante'].ru if fila['estudiante'].ru else "-"
        ]
        
        for idx_c, valor in enumerate(datos_identidad, 1):
            c = ws2.cell(row=i, column=idx_c, value=valor)
            c.alignment = align_center if idx_c in [1, 4, 5] else align_left
            c.fill = bg_color
            aplicar_borde(c)
        
        col_idx = 6
        for param in parametros_evaluacion:
            for act in param.actividades:
                c_nota = ws2.cell(row=i, column=col_idx, value=fila['detalle_actividades'][act.id])
                c_nota.alignment = align_center
                c_nota.fill = bg_color
                aplicar_borde(c_nota)
                col_idx += 1
            
            c_tot_cat = ws2.cell(row=i, column=col_idx, value=fila['notas_regulares'][param.id])
            c_tot_cat.alignment = align_center
            c_tot_cat.font = font_total_cat
            c_tot_cat.fill = fill_total_cat
            aplicar_borde(c_tot_cat)
            col_idx += 1
                
        if param_liberacion:
            for act in param_liberacion.actividades:
                c_lib = ws2.cell(row=i, column=col_idx, value=fila['detalle_actividades'][act.id])
                c_lib.alignment = align_center
                c_lib.fill = bg_color
                aplicar_borde(c_lib)
                col_idx += 1
                
            nota_lib = fila['nota_liberacion'] if fila['nota_liberacion'] and fila['nota_liberacion'] > 0 else "-"
            c_tot_lib = ws2.cell(row=i, column=col_idx, value=nota_lib)
            c_tot_lib.alignment = align_center
            c_tot_lib.font = font_bold
            c_tot_lib.fill = fill_liberacion if nota_lib != "-" else bg_color
            aplicar_borde(c_tot_lib)
            col_idx += 1
            
        c_sem = ws2.cell(row=i, column=col_idx, value=fila['nota_semestre'])
        c_sem.alignment = align_center
        c_sem.fill = fill_semestre
        aplicar_borde(c_sem)
        col_idx += 1
        
        if param_liberacion:
            c_lib_fin = ws2.cell(row=i, column=col_idx, value=nota_lib)
            c_lib_fin.alignment = align_center
            c_lib_fin.fill = fill_liberacion
            aplicar_borde(c_lib_fin)
            col_idx += 1
            
        c_fin = ws2.cell(row=i, column=col_idx, value=fila['nota_final'])
        c_fin.alignment = align_center
        c_fin.font = font_bold
        c_fin.fill = fill_final
        aplicar_borde(c_fin)

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    nombre_archivo = f"Notas_{paralelo.materia.sigla}_{paralelo.nombre}_Completo.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre_archivo
    )